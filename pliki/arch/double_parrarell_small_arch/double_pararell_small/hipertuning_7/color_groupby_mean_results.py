import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel file containing calculated averages
file_path = 'averages_results_7.xlsx'
df = pd.read_excel(file_path)

# Load the workbook using openpyxl to apply formatting
wb = load_workbook(file_path)
ws = wb.active

# Find the highest average for each group
highlight_indices = df.loc[df.groupby('Column')['Average'].idxmax()]

# Load the workbook using openpyxl to apply formatting
wb = load_workbook(file_path)
ws = wb.active

# Define the highlight fill
highlight_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# Apply the highlight to the entire row for each highest average
for idx, row in highlight_indices.iterrows():
    unique_value = row['Unique Value']
    col_name = row['Column']
    
    # Iterate over all rows in the sheet to find the row that matches the criteria
    for excel_row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        if excel_row[0].value == col_name and excel_row[1].value == unique_value:
            # Highlight the entire row
            for cell in excel_row:
                cell.fill = highlight_fill

# Save the workbook with highlights
wb.save('highlighted_averages.xlsx')
